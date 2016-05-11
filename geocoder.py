import openpyxl
import googlemaps, pprint

gmaps = googlemaps.Client(key='AIza...') #your API key goes here
woB = 'output.xlsx'
range = 'O2:O33246'
longLatRange = 'P2:Q33246'
resultDict = {}
wb = openpyxl.load_workbook(woB)
ws = wb.active

#imports only unique addresses and puts them in a list
def importToList(range):
    interList = []

    for row in ws.iter_rows(range):
        for cell in row:
            if cell.value not in interList: #only add values not already in list
                interList.append(cell.value)
    print('Unique Addresses to Geocode:', len(interList))
    return interList


#sends a geocode request to google and returns the input address + long + lat in list format
def req(address):
    result = gmaps.geocode(address)
    long = result[0].get('geometry').get('location').get('lng')
    lat = result[0].get('geometry').get('location').get('lat')
    r = [address,lat,long]
    return r

#creates a dictionary of addresses and long lats for quick output
def dictLyfe(aList):
    count = 0
    for ent in aList:
        output = req(ent)
        count += 1
        resultDict[output[0]] = output[1:]
        print('addresses geocoded so far: ', count)
    return resultDict

#uses the dictionary to populate a new excel sheet, saves said sheet
def addToSheet(dict):
    startRow = 2
    testCell = 'O%i' %startRow
    print('testCell', testCell)
    currentAddress = ws[testCell].value

    for row in ws.iter_rows(longLatRange):
        #print 'currentAddress', currentAddress
        colPos = 0
        for cell in row:
            #print dict.get('%s' % (currentAddress))[colPos]
            try:
                val = dict.get('%s' % (currentAddress))[colPos]
                print('val: ', val, 'col', colPos)
                cell.value = val
            except TypeError:
                cell.value = 'address not queried yet...'
            colPos += 1
        startRow +=1
        testCell = 'O%i' %startRow
        currentAddress = ws[testCell].value
        #if startRow > 2200:
         #   wb.save('first_set_geocoded.xlsx')
    wb.save('geocoded.xlsx')
    print(startRow, 'rows geocoded :)')

#execution zone:

list = importToList(range)
aList = list[0:20]
brule = dictLyfe(list)
#pprint.pprint(brule)
addToSheet(brule)

#resultDict[list]

#result = gmaps.geocode(list[0],language = 'JSON')
